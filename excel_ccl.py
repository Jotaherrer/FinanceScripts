import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import requests
import pandas_datareader as pdr
import datetime as dt
import json
import os
import time
from openpyxl import Workbook, load_workbook
from pandas.tseries.offsets import Day, MonthEnd
from api_config import usuario, password
from api_iol_recursiva import clean_assets

plt.style.use('fivethirtyeight')

# stocks = ['AAPL', 'ABEV', 'ABT', 'AMD', 'AMZN', 'ARCO', 'AUY', 'AXP', 'AZN', 'BA','BABA', 'BBD', 
#           'BCS','C','CSCO','CVX','DESP', 'DIS','EBAY','ERJ','FB','GE','GGB','GILD', 'GLOB',
#           'GOLD','GOOGL', 'GSK','HMY','IBM','INTC','ITUB','JNJ','JPM','KO','MCD','MELI','MMM',
#           'MSFT','NFLX','NKE','NVDA','PBR','PFE','PG','QCOM','SNAP','T','TS','TRIP','TSLA','TWTR',
#           'V','VALE','WFC','WMT','X','XOM']

# cedears = ['AAPL.BA', 'ABEV.BA', 'ABT.BA', 'AMD.BA', 'AMZN.BA', 'ARCO.BA', 'AUY.BA', 'AXP.BA', 'AZN.BA', 'BA.BA','BABA.BA', 'BBD.BA', 
#           'BCS.BA','C.BA','CSCO.BA','CVX.BA','DESP.BA', 'DISN.BA','EBAY.BA','ERJ.BA','FB.BA','GE.BA','GGB.BA','GILD.BA', 'GLNT.BA',
#           'GOLD.BA','GOOGL.BA', 'GSK.BA','HMY.BA','IBM.BA','INTC.BA','ITUB.BA','JNJ.BA','JPM.BA','KO.BA','MCD.BA','MELI.BA','MMM.BA',
#           'MSFT.BA','NFLX.BA','NKE.BA','NVDA.BA','PBR.BA','PFE.BA','PG.BA','QCOM.BA','SNAP.BA','T.BA','TEN.BA','TRIP.BA','TSLA.BA','TWTR.BA',
#           'V.BA','VALE.BA','WFC.BA','WMT.BA','X.BA','XOM.BA']

def token_key(text):
    """
    Function that works out the received content to extract refreshed access token
    """
    content2 = str(text.split())
    beginning = content2.find('access_token":"') + int(15)
    end = content2.find('token_type') - int(3)
    access_token = content2[beginning:end]
    return access_token

def get_access():
    url1 = "https://api.invertironline.com/token"

    data = {
            "username": usuario,
            "password": password,
            "grant_type": "password"        
    }
    response = requests.post(url1, data=data)
    if response.status_code == 200:
        content = response.text
        access_key = token_key(content)

        url2 = f'https://api.invertironline.com/api/v2/Cotizaciones/Acciones/CEDEARs/argentina'
        datos = requests.get(url2, headers={
                                        'Authorization': 'Bearer '+access_key
        })
        datos = json.loads(datos.text)
        datos = datos['titulos']
        datos = clean_assets(datos)
    return datos

def prices(ticker):
    try:
        start = dt.datetime.today()
        start = start.strftime('%Y-%m-%d') 
        price = pdr.get_data_yahoo(ticker, start=start)['Adj Close'][0]
        vol = pdr.get_data_yahoo(ticker, start=start)['Volume'][0]
    except KeyError:
            start = dt.datetime.today()
            start = start - Day(1)
            start = start.strftime('%Y-%m-%d') 
            price = pdr.get_data_yahoo(ticker, start=start)['Adj Close'][0]
            vol = pdr.get_data_yahoo(ticker, start=start)['Volume'][0]
    return (price,vol)

def excel(ticker, p_usa, p_arg, f, ccl, desvio):
    if os.path.exists('CCL.xlsx'):
        wb = load_workbook(filename='CCL.xlsx')
        ws = wb['Cotizaciones']
        fecha = time.asctime()
        ws.append([ticker,p_usa, p_arg, fecha, f, ccl, desvio])
        wb.save(filename='CCL.xlsx')
        wb.close()
        print('Carga exitosa de datos.')
    else:
        wb = Workbook()
        ws = wb.active
        ws.title='Cotizaciones'
        fecha = time.asctime()
        ws.append(['Ticker', ' Px. USA', 'Px. ARG.', 'Horario', 'Factor', 'CCL', 'Desvio'])
        ws.append([ticker,p_usa, p_arg, fecha, f, ccl, desvio])
        wb.save(filename='CCL.xlsx')
        wb.close()
        print('Creacion exitosa del Excel.')

i = 0
while i < 10:
    spx_data = []
    arg_data = []
    stocks = ['AAPL', 'AMZN', 'GOLD', 'BABA', 'PBR', 'TSLA']
    cedears = get_access().loc[:,['Ticker', 'Price', 'Horario','Cant. Ops.', 'Spread']]
    factores = [('AAPL',10),('AMZN',72),('GOLD',1),('BABA',9),('PBR',1),('TSLA',15)]

    for stock in stocks:
        price = prices(stock)[0]
        volume = prices(stock)[1]
        spx_data.append((stock, price, volume))

    df_stocks = pd.DataFrame(spx_data, columns=['Ticker', 'precio_usa', 'volumen_usa'])
    df_fact = pd.DataFrame(factores, columns=['Ticker', 'factor'])

    comb = df_stocks.merge(cedears,how='left').merge(df_fact, how='left')
    combinado = comb.loc[:,['Ticker', 'precio_usa','Price','Horario','factor']]
    combinado['CCL'] = combinado['Price'] / combinado['precio_usa'] * combinado['factor']
    combinado.rename(columns={'precio_usa':'Px. USA', "Price": 'Px. ARG.', "factor": 'Factor'},inplace=True)
    combinado['Desvio'] = combinado.apply(lambda row: round(((row['CCL'] / np.mean(combinado['CCL']) - 1) * 100),2), axis=1)

    for e in range(len(combinado['Ticker'])):
        excel(combinado['Ticker'][e],combinado['Px. USA'][e], combinado['Px. ARG.'][e], combinado['Factor'][e], combinado['CCL'][e], combinado['Desvio'][e])
        
    i += 1
    